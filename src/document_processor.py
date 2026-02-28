"""
Document Processor — extract text and split into meaningful chunks.
Supported formats: PDF, DOCX, TXT, XLSX

Third-party libraries are imported lazily (inside functions) so missing
packages only raise an error when that specific format is actually used.
"""

import io
import re
from dataclasses import dataclass


@dataclass
class Chunk:
    index: int
    text: str
    source: str


# ---------------------------------------------------------------------------
# Text extraction (lazy imports)
# ---------------------------------------------------------------------------

def extract_text_from_pdf(file_bytes: bytes) -> str:
    import pdfplumber
    text_parts = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text.strip())
    return "\n\n".join(text_parts)


def extract_text_from_docx(file_bytes: bytes) -> str:
    import docx
    doc = docx.Document(io.BytesIO(file_bytes))
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def extract_text_from_txt(file_bytes: bytes) -> str:
    return file_bytes.decode("utf-8", errors="replace").strip()


def extract_text_from_xlsx(file_bytes: bytes) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"[Sheet: {sheet_name}]")
        for row in ws.iter_rows(values_only=True):
            row_values = [str(cell) for cell in row if cell is not None]
            if row_values:
                parts.append(" | ".join(row_values))
    return "\n".join(parts)


def extract_text(file_bytes: bytes, filename: str) -> str:
    ext = filename.rsplit(".", 1)[-1].lower()
    extractors = {
        "pdf":  extract_text_from_pdf,
        "docx": extract_text_from_docx,
        "txt":  extract_text_from_txt,
        "xlsx": extract_text_from_xlsx,
    }
    if ext not in extractors:
        raise ValueError(f"Unsupported file type: .{ext}")
    return extractors[ext](file_bytes)


# ---------------------------------------------------------------------------
# Chunking
# ---------------------------------------------------------------------------

CHUNK_SIZE = 800
CHUNK_OVERLAP = 100


def _split_into_sentences(text: str) -> list[str]:
    sentences = re.split(r'(?<=[.!?])\s+', text)
    return [s.strip() for s in sentences if s.strip()]


def chunk_text(text: str, filename: str) -> list[Chunk]:
    text = re.sub(r'\n{3,}', '\n\n', text.strip())
    paragraphs = text.split("\n\n")

    sentences: list[str] = []
    for para in paragraphs:
        sentences.extend(_split_into_sentences(para))
        sentences.append("")  # paragraph boundary marker

    chunks: list[Chunk] = []
    current: list[str] = []
    current_len = 0
    overlap_tail = ""

    for sentence in sentences:
        if not sentence:
            if current_len >= CHUNK_SIZE // 2:
                body = overlap_tail + " ".join(current)
                chunks.append(Chunk(index=len(chunks), text=body.strip(), source=filename))
                overlap_tail = body[-CHUNK_OVERLAP:]
                current, current_len = [], 0
            continue

        current.append(sentence)
        current_len += len(sentence)

        if current_len >= CHUNK_SIZE:
            body = overlap_tail + " ".join(current)
            chunks.append(Chunk(index=len(chunks), text=body.strip(), source=filename))
            overlap_tail = body[-CHUNK_OVERLAP:]
            current, current_len = [], 0

    if current:
        body = overlap_tail + " ".join(current)
        chunks.append(Chunk(index=len(chunks), text=body.strip(), source=filename))

    return [c for c in chunks if len(c.text) > 50]


def process_document(file_bytes: bytes, filename: str) -> list[Chunk]:
    raw_text = extract_text(file_bytes, filename)
    return chunk_text(raw_text, filename)
